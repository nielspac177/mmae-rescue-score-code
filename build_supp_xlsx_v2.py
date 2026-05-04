"""Build supp_tables.xlsx — multi-sheet supplementary tables for the manuscript."""
from __future__ import annotations
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
V2 = HERE / "v2"

NAVY = "374E55"
GOLD = "DF8F44"
LIGHT = "F5F2EA"
WHITE = "FFFFFF"


def style_sheet(ws, df, title=None, header_fill=NAVY, col_widths=None):
    if title:
        ws["A1"] = title
        ws["A1"].font = Font(name="Calibri", size=13, bold=True, color=NAVY)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(df.columns))
        start_row = 3
    else:
        start_row = 1

    # Header row
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=str(col))
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=header_fill)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color="000000"))

    # Body rows
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            v = row[col]
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = border
            if (i - start_row) % 2 == 0:
                c.fill = PatternFill("solid", fgColor=LIGHT)

    # Column widths
    if col_widths:
        for j, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    else:
        for j, col in enumerate(df.columns, start=1):
            maxlen = max([len(str(col))] + [len(str(v)) for v in df.iloc[:, j-1]])
            ws.column_dimensions[get_column_letter(j)].width = min(max(12, maxlen + 2), 50)

    # Header row height
    ws.row_dimensions[start_row].height = 30


def main():
    summary = json.loads((V2 / "summary_v2.json").read_text())
    s = summary
    table1 = pd.read_csv(V2 / "table1_baseline.csv")
    m1_tab = pd.read_csv(V2 / "m1_risk_by_score.csv")
    m2_tab = pd.read_csv(V2 / "m2_risk_by_score.csv")
    uni = pd.read_csv(V2 / "univariate_ors.csv")
    m1_co = pd.read_csv(V2 / "m1_logit_coefs.csv")
    m2_co = pd.read_csv(V2 / "m2_logit_coefs.csv")

    label_map = {
        "age_pts": "Age (per category)",
        "sdh_vol_ge100": "SDH volume ≥100 mL",
        "anticoag": "Anticoagulation",
        "no_focal_deficit": "Absence of focal deficit",
        "plt_lt150": "Platelets <150 ×10⁹/L",
        "antiplatelet": "Antiplatelet therapy",
        "ant_post": "Anterior + posterior embolization",
    }
    m1_co_n = m1_co[m1_co["variable"] != "const"].copy()
    m2_co_n = m2_co[m2_co["variable"] != "const"].copy()
    m1_co_n["variable"] = m1_co_n["variable"].map(label_map)
    m2_co_n["variable"] = m2_co_n["variable"].map(label_map)

    def fmt_score_tab(t):
        out = pd.DataFrame()
        out["Score"] = t["score"].astype(int)
        out["n"] = t["n"].astype(int)
        out["Failures"] = t["failures"].astype(int)
        out["Rate (%)"] = (t["rate"] * 100).round(1)
        out["95% CI lower (%)"] = (t["ci_lo"] * 100).round(1)
        out["95% CI upper (%)"] = (t["ci_hi"] * 100).round(1)
        return out

    def fmt_or_tab(df, col="variable"):
        out = pd.DataFrame()
        out["Variable"] = df[col]
        out["OR"] = df["OR"].round(2)
        out["95% CI lower"] = df["OR_lo"].round(2)
        out["95% CI upper"] = df["OR_hi"].round(2)
        out["P value"] = df["p"].apply(lambda x: "<0.001" if x < 0.001 else f"{x:.3f}")
        return out

    # Score components definitions
    score_def = pd.DataFrame([
        ("Age <65 years", 0, "Yes", "Yes"),
        ("Age 65–80 years", 1, "Yes", "Yes"),
        ("Age >80 years", 2, "Yes", "Yes"),
        ("SDH volume ≥100 mL", 1, "Yes", "Yes"),
        ("Anticoagulation", 1, "Yes", "Yes"),
        ("Absence of focal deficit", 1, "Yes", "Yes"),
        ("Platelets <150 ×10⁹/L", 1, "Yes", "No"),
        ("Antiplatelet therapy", 1, "Yes", "No"),
        ("Anterior + posterior embolization", 1, "Yes", "No"),
    ], columns=["Variable", "Points", "In Model 1 (full)", "In Model 2 (simple)"])

    # Model metrics
    metrics = pd.DataFrame([
        ("Sample size (n)", s["n"], s["n"]),
        ("Events (rescue surgery)", s["events"], s["events"]),
        ("Score AUC, apparent",
         round(s["m1_score_auc"]["apparent"], 3),
         round(s["m2_score_auc"]["apparent"], 3)),
        ("Score AUC, optimism-corrected (Harrell)",
         round(s["m1_score_auc"]["corrected"], 3),
         round(s["m2_score_auc"]["corrected"], 3)),
        ("Logit AUC, apparent",
         round(s["m1_logit"]["apparent"], 3),
         round(s["m2_logit"]["apparent"], 3)),
        ("Logit AUC, optimism-corrected",
         round(s["m1_logit"]["corrected"], 3),
         round(s["m2_logit"]["corrected"], 3)),
        ("Brier score (logit)",
         round(s["m1_logit"]["brier"], 3),
         round(s["m2_logit"]["brier"], 3)),
        ("Hosmer–Lemeshow χ²",
         round(s["m1_hl"]["chi2"], 2),
         round(s["m2_hl"]["chi2"], 2)),
        ("Hosmer–Lemeshow P",
         round(s["m1_hl"]["p"], 3),
         round(s["m2_hl"]["p"], 3)),
        ("Max possible score", 8, 5),
        ("Recommended cutoff", "≥5", "≥4"),
    ], columns=["Metric", "Model 1 (full, 8-pt)", "Model 2 (simple, 5-pt)"])

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1 — Table 1
    ws = wb.create_sheet("Table1_Baseline")
    style_sheet(ws, table1,
                title="Table 1. Baseline characteristics by rescue status (n=214)",
                col_widths=[42, 24, 24, 22, 12])

    # Sheet 2 — Score components
    ws = wb.create_sheet("ScoreComponents")
    style_sheet(ws, score_def,
                title="Table 2. Point definitions for Model 1 (full) and Model 2 (simple)",
                col_widths=[40, 12, 22, 24])

    # Sheet 3 — M1 risk by score
    ws = wb.create_sheet("M1_RiskByScore")
    style_sheet(ws, fmt_score_tab(m1_tab),
                title="Table 3. Model 1 (full, 8 pts) — observed rescue rate by total score",
                col_widths=[10, 10, 12, 14, 18, 18])

    # Sheet 4 — M2 risk by score
    ws = wb.create_sheet("M2_RiskByScore")
    style_sheet(ws, fmt_score_tab(m2_tab),
                title="Table 4. Model 2 (simple, 5 pts) — observed rescue rate by total score",
                col_widths=[10, 10, 12, 14, 18, 18], header_fill=GOLD)

    # Sheet 5 — Univariate ORs
    ws = wb.create_sheet("UnivariateORs")
    style_sheet(ws, fmt_or_tab(uni),
                title="Table 5. Univariable logistic-regression odds ratios",
                col_widths=[34, 12, 16, 16, 14])

    # Sheet 6 — M1 multivariable
    ws = wb.create_sheet("M1_MultivariableOR")
    style_sheet(ws, fmt_or_tab(m1_co_n),
                title="Table 6. Multivariable logistic regression — Model 1 (full)",
                col_widths=[34, 12, 16, 16, 14])

    # Sheet 7 — M2 multivariable
    ws = wb.create_sheet("M2_MultivariableOR")
    style_sheet(ws, fmt_or_tab(m2_co_n),
                title="Table 7. Multivariable logistic regression — Model 2 (simple)",
                col_widths=[34, 12, 16, 16, 14], header_fill=GOLD)

    # Sheet 8 — Model metrics
    ws = wb.create_sheet("ModelMetrics")
    style_sheet(ws, metrics,
                title="Table 8. Discrimination and calibration metrics",
                col_widths=[40, 22, 22])

    # Sheet 9 — ML comparison
    if (V2 / "ml_comparison.csv").exists():
        ml = pd.read_csv(V2 / "ml_comparison.csv")
        ml_f = pd.DataFrame()
        ml_f["Model"] = ml["model"]
        ml_f["AUC apparent"] = ml["apparent"].round(3)
        ml_f["95% CI lower"] = ml["ci_lo"].round(3)
        ml_f["95% CI upper"] = ml["ci_hi"].round(3)
        ml_f["Bootstrap mean"] = ml["boot_mean"].round(3)
        ws = wb.create_sheet("ML_Comparison")
        style_sheet(ws, ml_f,
                    title="Table 9. Machine-learning comparison (5x10 stratified CV; 1000-bootstrap 95% CIs)",
                    col_widths=[28, 14, 14, 14, 16])

    # Sheet 10 — Operating points
    if (V2 / "operating_points.csv").exists():
        op = pd.read_csv(V2 / "operating_points.csv")
        ws = wb.create_sheet("OperatingPoints")
        style_sheet(ws, op,
                    title="Table 10. Operating-point metrics at candidate cutoffs (sens/spec/PPV/NPV with Wilson 95% CIs)",
                    col_widths=[26, 10, 6, 6, 6, 6, 22, 22, 22, 22])

    # Sheet 9 — README
    ws = wb.create_sheet("README", 0)  # first sheet
    ws["A1"] = "Supplementary Tables — MMA embolization rescue risk score (v2)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=NAVY)
    ws.merge_cells("A1:C1")
    ws["A3"] = (
        "Single-center retrospective cohort of 214 consecutive patients undergoing "
        "MMA embolization for chronic subdural hematoma; 36 rescue events.")
    ws["A3"].font = Font(name="Calibri", size=10, italic=True, color="555555")
    ws.merge_cells("A3:C3")
    ws.row_dimensions[3].height = 32
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")

    contents = [
        ("Sheet", "Contents"),
        ("Table1_Baseline", "Baseline characteristics by rescue status, with P values."),
        ("ScoreComponents", "Point definitions for Model 1 (8-pt) and Model 2 (5-pt)."),
        ("M1_RiskByScore", "Observed rescue rate by total Model 1 score with Wilson 95% CIs."),
        ("M2_RiskByScore", "Observed rescue rate by total Model 2 score with Wilson 95% CIs."),
        ("UnivariateORs", "Univariable logistic-regression odds ratios."),
        ("M1_MultivariableOR", "Multivariable logistic regression for Model 1."),
        ("M2_MultivariableOR", "Multivariable logistic regression for Model 2."),
        ("ModelMetrics", "Discrimination (AUC), calibration (Brier, Hosmer–Lemeshow), and cutoffs."),
    ]
    for i, (k, v) in enumerate(contents, start=5):
        is_h = i == 5
        ws.cell(row=i, column=1, value=k).font = Font(
            name="Calibri", size=10, bold=is_h,
            color="FFFFFF" if is_h else "222222")
        ws.cell(row=i, column=2, value=v).font = Font(
            name="Calibri", size=10, bold=is_h,
            color="FFFFFF" if is_h else "222222")
        if is_h:
            for col in (1, 2):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=NAVY)
        else:
            for col in (1, 2):
                cell = ws.cell(row=i, column=col)
                cell.fill = PatternFill("solid",
                                         fgColor=LIGHT if (i - 5) % 2 == 0 else WHITE)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 80

    out = HERE / "supp_tables.xlsx"
    wb.save(out)
    print(f"Wrote {out.relative_to(HERE)} ({out.stat().st_size/1024:.0f} KB) — {len(wb.sheetnames)} sheets")
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
